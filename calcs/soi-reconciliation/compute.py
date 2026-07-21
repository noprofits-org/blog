"""
Reconciling our Form 990 extract cut against the IRS's own published SOI tables.

WHAT THIS CHECKS
----------------
Four posts on this blog (months-of-cash-at-scale, nobody's-average, below-zero,
public-support-cliff) compute sector aggregates from the SOI *annual extract* of
Form 990 financial data. Two of them -- nobody's-average and public-support-cliff
-- carry a disclaimer saying the figures have not been reconciled against the
IRS's own published SOI tables; nobody's-average adds that this "is the check
that would catch a processing error."

This script is that check.

All four posts select rows the same way: one return per EIN, keeping the latest
tax period (`dedup` below). So `dedup` is the cut this blog actually ships, and
the one whose agreement with the published table matters most.

THE TWO SOURCES ARE NOT THE SAME THING
--------------------------------------
  PUBLISHED  22eo01.xlsx -- SOI Table 1, "Form 990 Returns of 501(c)(3)
             Organizations: Balance Sheet and Income Statement Items, by Asset
             Size", tax year 2022. Figures are WEIGHTED ESTIMATES BASED ON A
             SAMPLE, keyed to a TAX YEAR, in THOUSANDS of dollars. Excludes
             private foundations, most churches, and certain other religious
             organizations, and most organizations with receipts under $50,000.
             Latest published year as of July 2026.

  OURS       23eoextract990.zip -- the SOI annual extract, an ADMINISTRATIVE
             POPULATION file of every Form 990 PROCESSED in CALENDAR YEAR 2023,
             in DOLLARS. Not a sample; not a tax year.

So they disagree by construction, before any pipeline error enters. The point of
this script is to decompose the disagreement, not to declare a pass or a fail.

WHY WE COMPARE THE CY2023 EXTRACT AND NOT THE CY2024 ONE
--------------------------------------------------------
The published table is TY2022. Returns for tax year 2022 are filed mostly during
calendar year 2023, so the CY2023 extract is the analogue of the published
table's sampling frame. Comparing our usual CY2024 file to a TY2022 table would
confound the pipeline question with a full year of sector growth.

THE THREE CUTS
--------------
  pooled   every 501(c)(3) row in the CY2023 file (261,146 returns)
  ty2022   rows whose tax period falls in tax year 2022 (221,549 returns)
  dedup    one return per EIN, keeping the latest tax period (245,065 returns)
           <- THIS IS THE CUT THE PUBLISHED POSTS USE

The headline finding is that `pooled` -- the least "clean" of the three -- tracks
the published dollar totals best; `dedup`, the cut we ship, runs 3 to 8 percent
low on dollar levels; and `ty2022`, the filter that looks most correct, is the
worst of the three on every line item. See the post for why.

Ratios are far more robust than levels: the contributions-share-of-revenue
figure this blog leans on is within a point of the published estimate under all
three cuts, because the dedup bias hits numerator and denominator alike.

SOURCES
  https://www.irs.gov/pub/irs-soi/22eo01.xlsx
  https://www.irs.gov/pub/irs-soi/23eoextract990.zip
Run from calcs/soi-reconciliation/.  Deterministic; no random seed needed.
"""

import json
import pathlib

import openpyxl
import pandas as pd

DATA = pathlib.Path(__file__).resolve().parent.parent / "data"
OUT = pathlib.Path(__file__).resolve().parent

FAILURES = []


def check(label, got, want, tol):
    """Assert a computed value matches what the post claims, within tol."""
    ok = abs(got - want) <= tol
    if not ok:
        FAILURES.append(f"{label}: got {got!r}, want {want!r} (tol {tol})")
    print(f"  [{'ok' if ok else 'FAIL'}] {label}: {got:,.2f} (post says {want:,.2f})")
    return got


# ---------------------------------------------------------------- published --
# SOI Table 1, TY2022, "Total" column (column 1). Published in THOUSANDS of
# dollars; converted to dollars here so both sides are in the same unit.

wb = openpyxl.load_workbook(DATA / "22eo01.xlsx", data_only=True)
ws = wb["Sheet1"]
rows = {}
for r in range(6, ws.max_row + 1):
    label = ws.cell(r, 1).value
    total = ws.cell(r, 2).value
    if label is not None and total is not None:
        rows[str(label).strip()] = total

PUB = {
    "n": rows["Number of returns"],
    "assets": rows["Total assets"] * 1_000,
    "liab": rows["Total liabilities"] * 1_000,
    "netassets": rows["Total net assets"] * 1_000,
    "rev": rows["Total revenue"] * 1_000,
    "contrib": rows["Total contributions, gifts, and grants"] * 1_000,
    "govgrants": rows["Government grants (contributions)"] * 1_000,
    "prgmrev": rows["Program service revenue"] * 1_000,
    "exp": rows["Total expenses"] * 1_000,
}

print("\nPUBLISHED (SOI Table 1, TY2022, sample estimates)")
check("published number of returns", PUB["n"], 248_791, 0)
check("published total revenue ($B)", PUB["rev"] / 1e9, 3_028.0, 0.1)
check("published total assets ($B)", PUB["assets"] / 1e9, 6_174.8, 0.1)
check("published contributions ($B)", PUB["contrib"] / 1e9, 756.1, 0.1)
check("published total expenses ($B)", PUB["exp"] / 1e9, 2_909.4, 0.1)

# The one line the extract does not have at all: contributions split into
# government grants vs everything else. This is what settles the open question
# in the correction to nobody's-average.
gov_share = 100 * PUB["govgrants"] / PUB["contrib"]
check("government grants as % of contributions", gov_share, 40.3, 0.1)
check("contributions as % of revenue (published)", 100 * PUB["contrib"] / PUB["rev"], 24.97, 0.05)
check("program service revenue as % of revenue", 100 * PUB["prgmrev"] / PUB["rev"], 69.6, 0.1)

# --------------------------------------------------------------------- ours --

COLS = ["ein", "tax_pd", "subseccd", "totrevenue", "totassetsend",
        "totliabend", "totnetassetend", "totcntrbgfts", "totfuncexpns"]

df = pd.read_csv(DATA / "23eoextract990.zip", usecols=COLS, low_memory=False)
df["ty"] = df.tax_pd // 100
c3 = df[df.subseccd == 3].copy()


def agg(frame):
    return {
        "n": len(frame),
        "assets": frame.totassetsend.sum(),
        "liab": frame.totliabend.sum(),
        "netassets": frame.totnetassetend.sum(),
        "rev": frame.totrevenue.sum(),
        "contrib": frame.totcntrbgfts.sum(),
        "exp": frame.totfuncexpns.sum(),
    }


CUTS = {
    "pooled": c3,
    "ty2022": c3[c3.ty == 2022],
    "dedup": c3.sort_values("tax_pd").drop_duplicates("ein", keep="last"),
}
OURS = {name: agg(frame) for name, frame in CUTS.items()}

print("\nOUR CUTS (CY2023 extract, 501(c)(3) Form 990 filers)")
check("pooled returns", OURS["pooled"]["n"], 261_146, 0)
check("ty2022-only returns", OURS["ty2022"]["n"], 221_549, 0)
check("deduped returns", OURS["dedup"]["n"], 245_065, 0)
check("excess rows over organizations", OURS["pooled"]["n"] - OURS["dedup"]["n"], 16_081, 0)

# Distinguish the excess-row count from the number of rows actually involved in
# a repeat, which is larger because some EINs appear more than twice.
vc = c3.ein.value_counts()
check("EINs appearing more than once", (vc > 1).sum(), 14_279, 0)
check("rows involved in a repeat", vc[vc > 1].sum(), 30_360, 0)
check("largest number of rows for one EIN", vc.max(), 14, 0)

# Every duplicate EIN is a distinct tax period -- organizations catching up on a
# late filing, not the same return counted twice.
assert len(c3) == len(c3.drop_duplicates(["ein", "tax_pd"])), \
    "found a genuine duplicate (same EIN and same tax period)"

# ----------------------------------------------------------------- the gaps --

print("\nGAPS: (ours - published) / published, in percent")
GAPS = {}
for name, got in OURS.items():
    GAPS[name] = {k: 100 * (got[k] - PUB[k]) / PUB[k] for k in got}
    line = "  ".join(f"{k}={GAPS[name][k]:+6.2f}%" for k in ("n", "rev", "assets", "exp"))
    print(f"  {name:8s} {line}")

# The headline: pooled is closest on dollars, ty2022 is worst on everything.
check("pooled revenue gap (%)", GAPS["pooled"]["rev"], -0.08, 0.02)
check("pooled net assets gap (%)", GAPS["pooled"]["netassets"], -0.97, 0.02)
check("pooled assets gap (%)", GAPS["pooled"]["assets"], -1.92, 0.02)
check("pooled contributions gap (%)", GAPS["pooled"]["contrib"], +3.50, 0.02)
check("pooled expenses gap (%)", GAPS["pooled"]["exp"], -3.92, 0.02)
check("pooled count gap (%)", GAPS["pooled"]["n"], +4.97, 0.02)

check("ty2022 count gap (%)", GAPS["ty2022"]["n"], -10.95, 0.02)
check("ty2022 assets gap (%)", GAPS["ty2022"]["assets"], -12.04, 0.02)
check("ty2022 revenue gap (%)", GAPS["ty2022"]["rev"], -6.91, 0.02)

check("dedup count gap (%)", GAPS["dedup"]["n"], -1.50, 0.02)
check("dedup assets gap (%)", GAPS["dedup"]["assets"], -7.57, 0.02)
check("dedup contributions gap (%)", GAPS["dedup"]["contrib"], -0.84, 0.02)
check("dedup net assets gap (%)", GAPS["dedup"]["netassets"], -7.12, 0.02)
check("dedup expenses gap (%)", GAPS["dedup"]["exp"], -6.89, 0.02)
check("dedup revenue gap (%)", GAPS["dedup"]["rev"], -3.59, 0.02)

# Worst-case absolute gap across the seven line items, per cut. This is the
# number the post uses to rank the three cuts.
for name in CUTS:
    worst = max(abs(v) for v in GAPS[name].values())
    print(f"  worst-case gap, {name:8s} = {worst:5.2f}%")
check("pooled worst-case gap (%)", max(abs(v) for v in GAPS["pooled"].values()), 4.97, 0.02)
check("ty2022 worst-case gap (%)", max(abs(v) for v in GAPS["ty2022"].values()), 13.35, 0.02)
check("dedup worst-case gap (%)", max(abs(v) for v in GAPS["dedup"].values()), 8.37, 0.02)

# -------------------------------------------- the components we cannot see --
# Figure 2 in the post. The extract has only the contributions total; the
# published table splits it six ways. Values in billions of dollars.

COMPONENTS = {
    "Government grants (contributions)": 305.1,
    "All other contributions, gifts, grants, etc.": 396.1,
    "Related organizations": 37.8,
    "Fundraising events": 10.7,
    "Membership dues": 4.5,
    "Federated campaigns": 1.9,
}
print("\nCONTRIBUTION COMPONENTS (published, $B)")
for label, want in COMPONENTS.items():
    check(label, rows[label] / 1e6, want, 0.05)
minor = sum(rows[k] for k in list(COMPONENTS)[2:]) / 1e6
check("four minor components combined ($B)", minor, 55.0, 0.05)
check("six components sum to the total ($B)",
      sum(rows[k] for k in COMPONENTS) / 1e6, PUB["contrib"] / 1e9, 0.05)

# ------------------------------------------------- does any of this move us? --
# The live question for the published posts: the headline aggregate share of
# contributions in revenue. Published TY2022 says 24.97%. The published posts
# report 24.4% from a deduped CY2024 cut. The like-for-like figure here is the
# deduped CY2023 cut.

shares = {k: 100 * OURS[k]["contrib"] / OURS[k]["rev"] for k in OURS}
check("dedup contributions share (%)", shares["dedup"], 25.68, 0.05)
check("pooled contributions share (%)", shares["pooled"], 25.87, 0.05)
check("ty2022 contributions share (%)", shares["ty2022"], 25.15, 0.05)

pub_share = 100 * PUB["contrib"] / PUB["rev"]
print(f"\n  published TY2022 share  = {pub_share:.2f}%")
for k, v in shares.items():
    print(f"  our CY2023 {k:7s} share = {v:.2f}%  ({v - pub_share:+.2f} pts)")

# The point: every cut lands within ~1 point on the RATIO, even the cuts that
# are 8-13% off on LEVELS. Ratios survive what levels do not.
spread = max(abs(v - pub_share) for v in shares.values())
check("worst ratio error across all cuts (pts)", spread, 0.90, 0.02)
print("  -> the 24-25% headline in nobody's average survives the check.")

# ------------------------------------------------------------------- output --

payload = {
    "published": PUB,
    "ours": OURS,
    "gaps": GAPS,
    "gov_share_of_contributions": gov_share,
    "published_contrib_share": pub_share,
    "ours_contrib_shares": shares,
    "components_billions": {k: rows[k] / 1e6 for k in COMPONENTS},
}
(OUT / "reconciliation.json").write_text(json.dumps(payload, indent=2, default=float))

print("\n" + ("ALL CHECKS PASSED" if not FAILURES else f"{len(FAILURES)} FAILURES"))
for f in FAILURES:
    print("  " + f)
assert not FAILURES, FAILURES
