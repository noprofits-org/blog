#!/usr/bin/env python3
"""
Government grants to 501(c)(3) organizations, by asset size and code section.

Supports posts/2026-07-22-two-fifths-government.md on blog.noprofits.org.
This script is published for download alongside the post; it is meant to be
read, disagreed with, and rerun.

WHAT IT DOES
------------
The reconciliation post (posts/2026-07-21-checking-our-work.md) established
that government grants are $305.1B of $756.1B in 501(c)(3) contributions --
40.3 percent -- using one cell of the IRS's published SOI Table 1 for tax year
2022. This script reads the rest of that table, plus Table 3, and asks WHO that
money goes to:

  * by ASSET SIZE  (Table 1, six size classes) -- government grants as a share
    of each class's contributions and of its total revenue, and each class's
    share of the $305B;
  * by CODE SECTION (Table 3, 501(c)(3) through 501(c)(9)) -- how completely
    government grant money concentrates in 501(c)(3).

WHAT IT CANNOT DO
-----------------
The "Government grants (contributions)" line is Form 990 Part VIII line 1e:
money a government GIVES. Money a government PAYS -- Medicare, Medicaid,
service contracts where the government is the direct payer -- is program
service revenue (line 2), and these published tables do not split program
service revenue by payer. (Hospital filers report some Medicaid-related
detail on Schedule H; that schedule feeds none of the tables used here.)
Per the Form 990 instructions, a government payment made
primarily for the direct benefit of the public is a contribution (line 1e),
while one primarily serving the payor's own needs is program service revenue;
per the Schedule A instructions (citing Rev. Rul. 83-153), Medicare and
Medicaid payments are gross receipts from patients, not government support.
So every figure here is a floor on government money, not a measure of it.

Also: the SOI annual extract (the per-filer file behind this blog's other
posts) does not carry line 1e at all, so none of this can be computed for any
individual organization from that file. For one organization, read its own
Form 990 Part VIII.

DATA (both cached in ../data/, both free from the IRS)
------------------------------------------------------
  22eo01.xlsx  SOI Table 1, "Form 990 Returns of 501(c)(3) Organizations:
               Balance Sheet and Income Statement Items, by Asset Size",
               tax year 2022. Weighted estimates from a sample, in THOUSANDS
               of dollars. Excludes private foundations, most churches, most
               orgs with receipts under $50k.
               https://www.irs.gov/pub/irs-soi/22eo01.xlsx
  22eo03.xlsx  SOI Table 3, same items for 501(c)(3)-(9) organizations, by
               code section. Same vintage, units, and caveats.
               https://www.irs.gov/pub/irs-soi/22eo03.xlsx

The small-asset-size columns rest on fewer sampled returns and the published
table flags some cells as "use with caution"; treat the sub-$1M columns as
noisier than the rest.

RUN
---
  python3 compute.py            # needs openpyxl
Deterministic; no random seed needed. Writes results.json for figures.py and
prints a check line for every IRS-derived number quoted in the post. (Survey
figures quoted from the Urban Institute are cited, not ingested, and are not
checked here.)
"""

import json
import pathlib

import openpyxl

DATA = pathlib.Path(__file__).resolve().parent.parent / "data"
OUT = pathlib.Path(__file__).resolve().parent

FAILURES = []


def check(label, got, want, tol):
    """Assert a computed value matches what the post claims, within tol."""
    ok = abs(got - want) <= tol
    if not ok:
        FAILURES.append(f"{label}: got {got!r}, want {want!r} (tol {tol})")
    print(f"  [{'ok' if ok else 'FAIL'}] {label}: {got:,.2f} (post says {want:,.2f})")
    return got


def read_table(path, ncols):
    """Row label -> list of column values (col 2 .. col 1+ncols), from row 6."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Sheet1"]
    rows = {}
    for r in range(6, ws.max_row + 1):
        lab = ws.cell(r, 1).value
        if lab is None:
            continue
        vals = [ws.cell(r, c).value for c in range(2, 2 + ncols)]
        if any(v is not None for v in vals):
            rows[str(lab).strip()] = vals
    return rows


# -------------------------------------------------- Table 1: by asset size --
# Columns: Total, <$100k, $100k-500k, $500k-1M, $1M-10M, $10M-50M, >=$50M.
# Published in thousands of dollars; ratios are unit-free, dollar figures are
# converted to $B where quoted.

t1 = read_table(DATA / "22eo01.xlsx", 7)
SIZES = ["total", "lt100k", "100k_500k", "500k_1m", "1m_10m", "10m_50m", "ge50m"]

N = t1["Number of returns"]
REV = t1["Total revenue"]
CON = t1["Total contributions, gifts, and grants"]
GOV = t1["Government grants (contributions)"]
PRG = t1["Program service revenue"]

print("\nHEADLINE (SOI Table 1, TY2022, sample estimates; same cells as the")
print("reconciliation post)")
check("returns", N[0], 248_791, 0)
check("total revenue ($B)", REV[0] / 1e6, 3_028.0, 0.1)
check("contributions ($B)", CON[0] / 1e6, 756.1, 0.1)
check("government grants ($B)", GOV[0] / 1e6, 305.1, 0.1)
check("gov grants as % of contributions", 100 * GOV[0] / CON[0], 40.3, 0.05)
check("gov grants as % of total revenue", 100 * GOV[0] / REV[0], 10.1, 0.05)
check("program service as % of revenue", 100 * PRG[0] / REV[0], 69.6, 0.05)

print("\nBY ASSET SIZE -- returns per class (Table 1 in the post)")
for i, want in enumerate([None, 30_333, 62_017, 37_838, 83_692, 24_040, 10_871]):
    if want is not None:
        check(f"  {SIZES[i]} returns", N[i], want, 0)

print("\nBY ASSET SIZE -- gov grants as % of the class's contributions")
pct_contrib = [100 * GOV[i] / CON[i] for i in range(7)]
for i, want in enumerate([None, 23.8, 26.6, 25.7, 41.0, 51.6, 37.9]):
    if want is not None:
        check(f"  {SIZES[i]}", pct_contrib[i], want, 0.05)

print("\nBY ASSET SIZE -- gov grants as % of the class's total revenue")
pct_rev = [100 * GOV[i] / REV[i] for i in range(7)]
for i, want in enumerate([None, 14.8, 17.1, 16.9, 23.4, 23.0, 7.0]):
    if want is not None:
        check(f"  {SIZES[i]}", pct_rev[i], want, 0.05)

print("\nBY ASSET SIZE -- share of the $305B, and dollars")
share_gov = [100 * GOV[i] / GOV[0] for i in range(7)]
gov_b = [GOV[i] / 1e6 for i in range(7)]
for i, (s_want, b_want) in enumerate(
    [(None, None), (0.5, 1.6), (1.6, 4.8), (1.4, 4.1),
     (16.3, 49.8), (24.5, 74.8), (55.7, 170.0)]
):
    if s_want is not None:
        check(f"  {SIZES[i]} share of gov $", share_gov[i], s_want, 0.05)
        check(f"  {SIZES[i]} gov $B", gov_b[i], b_want, 0.05)

check("top two classes' share of gov $", share_gov[5] + share_gov[6], 80.2, 0.05)
check("orgs under $1M assets", sum(N[1:4]), 130_188, 0)
check("  ... as % of returns", 100 * sum(N[1:4]) / N[0], 52.3, 0.05)
check("  ... their share of gov $", 100 * sum(GOV[1:4]) / GOV[0], 3.4, 0.05)
check(">=50M program service % of its revenue", 100 * PRG[6] / REV[6], 75.8, 0.05)
check(">=50M count", N[6], 10_871, 0)
check(">=50M share of returns (%)", 100 * N[6] / N[0], 4.4, 0.05)

# "Everything else" residual per class (investment income, royalties, net
# gains, etc.) for the composition figure; all residuals must be positive.
oth_pct = [100 * (REV[i] - CON[i] - PRG[i]) / REV[i] for i in range(7)]
assert all(v > 0 for v in oth_pct), "negative residual revenue component"
priv_pct = [100 * (CON[i] - GOV[i]) / REV[i] for i in range(7)]
prg_pct = [100 * PRG[i] / REV[i] for i in range(7)]
check("sector: everything-else % of revenue", oth_pct[0], 5.4, 0.05)
check("sector: non-gov contributions % of revenue", priv_pct[0], 14.9, 0.05)

# ------------------------------------------------ Table 3: by code section --
t3 = read_table(DATA / "22eo03.xlsx", 8)
SECS = ["total", "c3", "c4", "c5", "c6", "c7", "c8", "c9"]
REV3 = t3["Total revenue"]
GOV3 = t3["Government grants (contributions)"]

print("\nBY CODE SECTION (SOI Table 3, TY2022)")
check("all-sections gov grants ($B)", GOV3[0] / 1e6, 312.9, 0.05)
check("c3 share of all gov grants (%)", 100 * GOV3[1] / GOV3[0], 97.5, 0.05)
sec_pct_rev = [100 * GOV3[i] / REV3[i] for i in range(8)]
for i, want in enumerate([None, 10.1, 2.6, 1.0, 5.9, 0.3, None, None]):
    if want is not None:
        check(f"  {SECS[i]} gov % of own revenue", sec_pct_rev[i], want, 0.05)
check("c4 gov grants ($B)", GOV3[2] / 1e6, 4.1, 0.05)
check("c6 gov grants ($B)", GOV3[4] / 1e6, 3.5, 0.05)
check("c9 gov % of own revenue", sec_pct_rev[7], 0.004, 0.002)

# Every remaining cell of the post's Table 2, so the script asserts the whole
# table, not just the cells discussed in prose.
for i, (name, rev_b, gov_b3) in enumerate(
    [(None, None, None), ("c3", 3_028.0, 305.1), ("c4", 156.4, 4.1),
     ("c5", 29.1, 0.3), ("c6", 58.5, 3.5), ("c7", 18.4, 0.1),
     ("c8", 21.6, 0.0), ("c9", 161.6, 0.0)]
):
    if name is not None:
        check(f"  {name} revenue ($B)", REV3[i] / 1e6, rev_b, 0.05)
        check(f"  {name} gov grants ($B, as printed)", GOV3[i] / 1e6, gov_b3, 0.05)
check("c8 gov % of own revenue (post: <0.1)", sec_pct_rev[6], 0.03, 0.05)

# ------------------------------------------------- prose ratio claims --------
print("\nPROSE CLAIMS")
assert sum(1 for v in pct_contrib[1:] if v > 50) == 1, \
    "the $10M-50M class should be the ONLY majority-government class"
assert pct_rev[6] < min(pct_rev[1:6]), "giants should be lowest on the revenue panel"
assert pct_rev[6] <= 0.5 * min(pct_rev[1:4]) + 0.01, "'half the small classes' figure'"
assert abs(pct_rev[6] / max(pct_rev[4:6]) - 1 / 3) < 0.04, "'a third of the middle's'"
assert all(23 <= v <= 27 for v in pct_contrib[1:4]), "'about a quarter' (sub-$1M classes)"
assert sec_pct_rev[4] > 2 * sec_pct_rev[2], "c6 'more than double' c4's rate"
for i in range(7):
    assert abs(pct_rev[i] + priv_pct[i] + prg_pct[i] + oth_pct[i] - 100) < 1e-9, \
        "Figure 2 segments must sum to 100"
print("  [ok] all prose ratio claims hold; Figure 2 segments sum to 100")

# -------------------------------------------------------------------- save --
results = {
    "sizes": SIZES,
    "size_labels": ["Total", "<$100k", "$100k–500k", "$500k–1M",
                    "$1M–10M", "$10M–50M", "≥$50M"],
    "n": N, "rev_k": REV, "contrib_k": CON, "gov_k": GOV, "prgm_k": PRG,
    "pct_contrib": pct_contrib, "pct_rev": pct_rev,
    "share_gov": share_gov, "gov_b": gov_b,
    "prg_pct": prg_pct, "priv_pct": priv_pct, "oth_pct": oth_pct,
    "sections": SECS, "sec_rev_k": REV3, "sec_gov_k": GOV3,
    "sec_pct_rev": sec_pct_rev,
}
(OUT / "results.json").write_text(json.dumps(results, indent=1))
print(f"\nwrote {OUT / 'results.json'}")

if FAILURES:
    raise SystemExit("\nFAILED CHECKS:\n" + "\n".join(FAILURES))
print("all checks passed")
